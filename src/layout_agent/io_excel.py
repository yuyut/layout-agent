"""Load cube specifications, media requirements, and material flow from Excel."""
from __future__ import annotations

import logging
from pathlib import Path

from .models import Cube, FlowEdge

logger = logging.getLogger(__name__)


def load_problem_excel(xlsx_path: str | Path) -> tuple[list[Cube], list[FlowEdge]]:
    """
    Read the assignment Excel workbook and return (cubes, flows).

    Expected sheets
    ───────────────
    Characteristics  – columns: name | xDim | yDim | color
    Media            – columns: cubeName | water | electricity
                       'x' (case-insensitive) means required; blank means not.
    Materialflow     – columns: cube1 | cube2 | intensity  (sparse edge list)
                       Row 0 is the header; data starts at row 1.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required: pip install openpyxl")

    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)

    # ── Characteristics ──────────────────────────────────────────────────────
    if "Characteristics" not in wb.sheetnames:
        raise ValueError("Sheet 'Characteristics' not found in workbook.")

    ws_char = wb["Characteristics"]
    char_rows = list(ws_char.iter_rows(values_only=True))
    cubes_raw: dict[str, dict] = {}
    for row in char_rows[1:]:          # skip header
        if row[0] is None:
            continue
        cubes_raw[str(row[0])] = {
            "xdim":  float(row[1]),
            "ydim":  float(row[2]),
            "color": str(row[3]) if row[3] is not None else None,
        }
    logger.info(f"Characteristics: {len(cubes_raw)} cubes found")

    # ── Media ────────────────────────────────────────────────────────────────
    media: dict[str, dict[str, bool]] = {}
    if "Media" in wb.sheetnames:
        ws_media = wb["Media"]
        for row in list(ws_media.iter_rows(values_only=True))[1:]:
            if row[0] is None:
                continue
            media[str(row[0])] = {
                "needs_water":       (str(row[1]).strip().lower() == "x") if row[1] else False,
                "needs_electricity": (str(row[2]).strip().lower() == "x") if row[2] else False,
            }
        logger.info(
            f"Media: {sum(v['needs_water'] for v in media.values())} water, "
            f"{sum(v['needs_electricity'] for v in media.values())} electricity"
        )
    else:
        logger.warning("Sheet 'Media' not found; assuming no utility requirements.")

    # ── Assemble Cube objects ────────────────────────────────────────────────
    cubes: list[Cube] = []
    for cube_id, props in cubes_raw.items():
        m = media.get(cube_id, {"needs_water": False, "needs_electricity": False})
        cubes.append(Cube(
            id=cube_id,
            xdim=props["xdim"],
            ydim=props["ydim"],
            color=props["color"],
            needs_water=m["needs_water"],
            needs_electricity=m["needs_electricity"],
        ))

    # ── Materialflow ─────────────────────────────────────────────────────────
    flows: list[FlowEdge] = []
    if "Materialflow" in wb.sheetnames:
        ws_flow = wb["Materialflow"]
        flow_rows = list(ws_flow.iter_rows(values_only=True))
        for row in flow_rows[1:]:       # skip header row
            if row[0] is None or row[1] is None or row[2] is None:
                continue
            src, dst = str(row[0]), str(row[1])
            try:
                intensity = float(row[2])
            except (TypeError, ValueError):
                continue
            flows.append(FlowEdge(src=src, dst=dst, intensity=intensity))
        logger.info(f"Materialflow: {len(flows)} edges loaded")
    else:
        logger.warning("Sheet 'Materialflow' not found; no flow edges loaded.")

    return cubes, flows
